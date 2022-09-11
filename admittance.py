import csv
import os
from collections import defaultdict
from typing import List, Dict, Iterable, Optional, Set, Union, DefaultDict

import openpyxl as xl

from form_data import Person, Registration
from datetime import datetime


def _normalise(field: str) -> str:
    return field.lower().strip()


def read_registrations(file_path: str) -> List[Registration]:
    with open(file_path, encoding="utf-8") as registration_file:
        next(reader := csv.reader(registration_file))  # assign reader and skip headers
        return [
            Registration(
                _normalise(name),
                _normalise(mail),
                datetime.strptime(timestamp, '%d/%m/%Y %H:%M:%S'),
                [timeslot.replace(' ', '') for timeslot in timeslots.split(',')])
            for timestamp, mail, name, timeslots, *_ in reader
        ]


def read_people_table(file_path: str, name_column: int, email_column: int, allow_failure: bool = True) -> List[Person]:
    try:
        with open(file_path, encoding="utf-8") as person_table_file:
            next(table_reader := csv.reader(person_table_file))  # skip headers and assign to table_reader
            return [Person(_normalise(row[name_column]), _normalise(row[email_column])) for row in table_reader]
    except FileNotFoundError as file_not_found_error:
        if allow_failure:
            return []
        raise file_not_found_error


class Timeslot:
    spots: List[Person]
    disallowed: Set[Person]  # The people who are not allowed to attend this timeslot

    def __init__(self, disallowed: Optional[Set[Person]] = None):
        self.spots = []
        self.disallowed = disallowed or set()

    @property
    def spots_taken(self):
        return len(self.spots)

    def __str__(self):
        content = ', '.join(f"{k}: {str(v)}" for k, v in self.__dict__.items())
        return f"Timeslot({content})"

    def __repr__(self):
        content = ', '.join(f"{k}: {str(v)}" for k, v in self.__dict__.items())
        return f"Timeslot({content})"

    def admit(self, person: Person) -> bool:
        if person in self.disallowed:
            return False
        self.spots.append(person)
        return True

    def remove(self, person: Person) -> bool:
        try:
            self.spots.remove(person)
            return True
        except ValueError:
            return False


class LimitedTimeslot(Timeslot):
    capacity: int

    def __init__(self, capacity: int):
        super().__init__()
        self.capacity = capacity

    @property
    def spots_available(self):
        return self.capacity - self.spots_taken

    def admit(self, person: Person) -> bool:
        if person in self.disallowed:
            return False
        if self.spots_available > 0:
            self.spots.append(person)
            return True
        return False


class OpeningAdmittance:
    timeslots: Dict[str, Timeslot]
    processed: Dict[Person, Registration]
    cancelled: Set[Person]
    banned: Set[Person]
    marked: DefaultDict[Person, List[str]]
    confirmed_duplicates: Set[Person]

    def __init__(self, timeslots: Optional[Dict[str, Timeslot]] = None):
        self.timeslots = timeslots if timeslots else {}
        self.waiting_list = []
        self.cancelled = set()
        self.banned = set()
        self.marked = defaultdict(list)
        self.confirmed_duplicates = set()

    def clear(self):
        for timeslot in self.timeslots.values():
            timeslot.spots.clear()
        self.processed.clear()
        self.cancelled.clear()
        self.banned.clear()
        self.marked.clear()

    def _preprocess_and_mark(self, registrations: Iterable[Registration]):
        """
        Look through all regisstrations beforehand to mark individuals for manual checking if needed
        and overwrite duplicate registrations by the latest entry from said person if the latest entry makes changes
        to their preferences in timeslot
        :param registrations: All entries from the registration form
        :return: registrations without duplicate entries (NOTE: suspected duplicates are only marked for manual check
                 and will remain as separate registrations)
        """
        bad_email_endings = [".con", "@ntnu.no"]

        proccessed_for_admission = {}
        for registration in registrations:

            # Evaluate if peron is banned
            if registration.person in self.banned:
                self.marked[registration.person].append("Banned from attending, see ban list!")
                continue
            else:
                confirmed_duplicate = False
                for banned_person in self.banned:
                    if banned_person.similar(registration.person):
                        if confirmed_duplicate := registration.person in self.confirmed_duplicates:
                            self.marked[registration.person].append(f"Confirmed ban, see ban list for {banned_person}!")
                            self.banned.add(registration.person)
                            break
                        else:
                            self.marked[registration.person].append(f"Suspected ban: {registration.person} might be!, "
                                                                    f"{banned_person} from banlist!")
                            break
                if confirmed_duplicate:
                    continue  # skip this person, go on to the next!

            # Evaluate if person has a bad email ending

            for ending in bad_email_endings:
                if registration.person.email.endswith(ending):
                    self.marked[registration.person].append(f"Likely a non-working email! It ends with '{ending}'.")

            # Evaluate if person has not been given a timeslot because of attending previous "premium" timeslots in
            # earlier opening

            if all(registration.person in timeslot.disallowed for timeslot in self.timeslots.values()):
                self.marked[registration.person].append(
                    "Down prioritised from attending the timeslot(s) they signed up for, "
                    "attended previous opening in the early slot(s)!"
                )
                continue  # go on to the next person!

            for timeslot_name, timeslot in self.timeslots.items():
                if timeslot_name not in registration.timeslots:  # we only care if they signed this timeslot
                    continue
                if registration.person in timeslot.disallowed:
                    self.marked[registration.person].append(
                        f"Down prioritised from attending {timeslot_name} because they "
                        f"attended previous opening in the early slot(s)!"
                    )
                    break
                else:
                    # confirmed_duplicate = False
                    for downprioritised_person in timeslot.disallowed:
                        if downprioritised_person.similar(registration.person):
                            if confirmed_duplicate := registration.person in self.confirmed_duplicates:
                                self.marked[registration.person].append(
                                    f"Down prioritised from attending {timeslot_name} because they "
                                    "attended previous opening in the early slot(s)!. confirmed suspected duplicate"
                                    f" of: {downprioritised_person} from downprioritised list!"
                                )
                                timeslot.disallowed.add(registration.person)
                                break
                            else:
                                self.marked[registration.person].append(
                                    f"Subject to being down prioritised from {timeslot_name}, "
                                    f"suspecting {registration.person} might be the"
                                    f"same as {downprioritised_person} from the down prioritised list!"
                                )
                                break
                    # if confirmed_duplicate:
                    #     continue  # skip this person, go on to the next!

            # Evaluate if person is already in the system
            if (person := registration.person) in proccessed_for_admission.keys():
                # only overwrite entry if change in timeslots
                if set(registration.timeslots) != set(proccessed_for_admission[person].timeslots):
                    # NOTE: changing your timeslots has its drawback - you're now later in the queue
                    reason = f"Duplicate Entry for {person}:\noverwriting {proccessed_for_admission[person]}...\n" \
                             f"timestamp changed from {proccessed_for_admission[person].timestamp} to {registration.timestamp}\n" \
                             f"changed timeslots from {proccessed_for_admission[person].timeslots} to {registration.timeslots}"
                    self.marked[person].append(reason)
                else:
                    # if no substantial change is made, don't reprocess the person. They did as intended the first
                    # time around and should not be punished for trying to make sure they registered.
                    continue
            else:
                for already_processed_person, already_processed_registration in proccessed_for_admission.copy().items():
                    if registration.person.similar(already_processed_person):
                        if confirmed_duplicate := registration.person in self.confirmed_duplicates:
                            # only overwrite entry if change in timeslots
                            self.marked[already_processed_person].append(
                                f"Confirmed suspected duplicate! {registration.person} is the same as {already_processed_person}!"
                                f"\nOverwriting {already_processed_registration} with {registration}...\n"
                            )
                            del proccessed_for_admission[already_processed_person]
                        else:
                            self.marked[already_processed_person].append(f"Suspected duplicate of {registration}")
                            self.marked[registration.person].append(
                                f"Suspected duplicate of {already_processed_registration}")
                            break

            proccessed_for_admission[person] = registration
        return proccessed_for_admission

    def auto_admit(self, registrations: Iterable[Registration]):
        self.processed = self._preprocess_and_mark(registrations)
        for registration in self.processed.values():
            if not any(self.timeslots[wanted_slot].admit(registration) for wanted_slot in registration.timeslots if
                       wanted_slot in self.timeslots):
                self.waiting_list.append(registration)

    # def cancel(self, cancelled: Union[Iterable[Person], Person]):
    #     if isinstance(cancelled, Person):
    #         cancelled = (cancelled,)  # make iterable
    #     for person in cancelled:
    #         removed = False
    #         for timeslot in self.timeslots.values():
    #             for other in timeslot.spots.copy():
    #                 if person == other:
    #                     removed = timeslot.remove(person)
    #                     self.cancelled.add(person)
    #                 elif person.similar(other):
    #                     self.marked[other] = f"{self.marked.get(other, '')} Might have cancelled! " \
    #                                           f"{person} cancelled, and {other} might be the same person."
    #         if not removed:
    #             warnings.warn(f"Unable to cancel for {person}! They were not found in timeslots!")
    #         # TODO: also look in waiting lists, they should be removed from it too!
    #
    # def ban(self, banned: Union[Iterable[Person], Person]):
    #     if isinstance(banned, Person):
    #         banned = (banned,)  # make iterable
    #     for person in banned:
    #         removed = False
    #         for timeslot in self.timeslots.values():
    #             for other in timeslot.spots.copy():
    #                 if person == other:
    #                     removed = timeslot.remove(person)
    #                     self.banned.add(person)
    #                 else:
    #                     if person.similar(other):
    #                         self.marked[other] = f"{self.marked.get(other, '')} Might have been banned! " \
    #                                               f"{person} is banned, and {other} might be the same person."
    #         if not removed:
    #             warnings.warn(f"Unable to ban {person}! They were not found in timeslots!")
    #         # TODO: also look in waiting lists, they should be banned from it too!

    def write_to_spreadsheets(self, destination: str):
        workbook = xl.Workbook()
        for i, (timeslot_name, timeslot) in enumerate(self.timeslots.items()):
            timeslot_name = timeslot_name.replace(':', '_')
            sheet = workbook.create_sheet(timeslot_name, i)
            sheet.append(["Timestamp", "Name", "Email", "Wanted Timeslots", "Remarks"])
            for j, registration in enumerate(self.processed[person] for person in timeslot.spots):
                sheet.append([
                    registration.timestamp,
                    registration.person.name,
                    registration.person.email,
                    ", ".join(registration.timeslots),
                    '\n'.join(self.marked.get(registration.person, []))
                ])

        waiting_list_sheet = workbook.create_sheet("Waiting List", len(self.timeslots))
        waiting_list_sheet.append(["Timestamp", "Name", "Email", "Wanted Timeslots", "Remarks"])
        for j, registration in enumerate(self.waiting_list):
            waiting_list_sheet.append([
                registration.timestamp,
                registration.person.name,
                registration.person.email,
                ", ".join(registration.timeslots),
                '\n'.join(self.marked.get(registration.person, []))
            ])
        cancelled_list_sheet = workbook.create_sheet("Cancelled", len(self.timeslots) + 1)
        cancelled_list_sheet.append(["Name", "Email", "Remarks"])
        for j, person in enumerate(self.cancelled):
            cancelled_list_sheet.append([
                person.name,
                person.email,
                '\n'.join(self.marked.get(person, []))
            ])
        banned_list_sheet = workbook.create_sheet("Banned", len(self.timeslots) + 2)
        banned_list_sheet.append(["Name", "Email", "Remarks"])
        for j, person in enumerate(self.banned):
            banned_list_sheet.append([
                person.name,
                person.email,
                '\n'.join(self.marked.get(person, []))
            ])
        marked_list_sheet = workbook.create_sheet("All Remarks", len(self.timeslots) + 3)
        marked_list_sheet.append(["Name", "Email", "Remarks"])
        for j, (person, remarks) in enumerate(self.marked.items()):
            marked_list_sheet.append([
                person.name,
                person.email,
                '\n'.join(remarks)
            ])
        workbook.remove(workbook["Sheet"])
        workbook.save(destination + f"/output_{datetime.now().strftime('%Y%m%d__%H_%M_%S')}.xlsx")

    def write_to_csv(self):
        os.makedirs("./output/", exist_ok=True)
        for name, timeslot in self.timeslots.items():
            with open(f"./output/{name.replace(':', '')}.csv", 'w', newline='', encoding="utf-8") as file:
                fields = timeslot.spots[0].__dict__.keys()
                writer = csv.DictWriter(file, fields)
                writer.writeheader()
                for registration in timeslot.spots:
                    writer.writerow({k: v for k, v in registration.__dict__.items() if k in fields})

        with open(f"./output/waiting_list.csv", 'w', newline='', encoding="utf-8") as file:
            fields = self.waiting_list[0].__dict__.keys()
            writer = csv.DictWriter(file, fields)
            writer.writeheader()
            for registration in self.waiting_list:
                writer.writerow({k: v for k, v in registration.__dict__.items() if k in fields})

        with open(f"./output/marked.csv", 'w', newline='', encoding="utf-8") as file:
            marked_list = [(reg.person, reason) for reg, reason in self.marked.items()]
            fields = [*marked_list[0][0].__dict__.keys(), "marked reason"]
            writer = csv.DictWriter(file, fields)
            writer.writeheader()
            for person, reason in marked_list:
                entry = {k: v for k, v in person.__dict__.items() if k in fields}
                entry["marked reason"] = reason
                writer.writerow(entry)
